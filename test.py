print("\n")

import requests
import bs4
import openpyxl


url = "http://books.toscrape.com/catalogue/page-1.html"

base_url =  "http://books.toscrape.com/catalogue/page-{}.html"

file = openpyxl.load_workbook('template.xlsx')
sheet = file.active

(i1, i2, i3, i4, i5) = (3,3,3,3,3)


for n in range(1, 5):

    scrape_url = base_url.format(n)
    re = requests.get(scrape_url)

    soup = bs4.BeautifulSoup(re.text, "lxml")
    prices = soup.select('.price_color') # get book price
    books = soup.select('.product_pod') # get title & cover_link

    for book in books:
        # 1 stars books
        
        if len(book.select('.star-rating.One')) != 0:
            sheet1 = file['1 stars']
        
            money = book.select('.price_color')[0].text[1:]
            title = book('img')[0]['alt']
            source = book('img')[0]['src']
            cover_link = "http://books.toscrape.com"+source[2:]

            price_cell = sheet1.cell(row = i1, column = 3)
            price_cell.value = money

            title_cell = sheet1.cell(row = i1, column = 2)
            title_cell.value = title

            cover_cell = sheet1.cell(row = i1, column = 4)
            cover_cell.value = cover_link
            i1+=1

        # 2 stars books
        elif len(book.select('.star-rating.Two')) != 0:
            sheet2 = file['2 stars']

            money = book.select('.price_color')[0].text[1:]
            title = book('img')[0]['alt']
            source = book('img')[0]['src']
            cover_link = "http://books.toscrape.com"+source[2:]

            price_cell = sheet2.cell(row = i2, column = 3)
            price_cell.value = money

            title_cell = sheet2.cell(row = i2, column = 2)
            title_cell.value = title

            cover_cell = sheet2.cell(row = i2, column = 4)
            cover_cell.value = cover_link
            i2+=1

        # 3 stars books
        elif len(book.select('.star-rating.Three')) != 0:
            sheet3 = file['3 stars']

            money = book.select('.price_color')[0].text[1:]
            title = book('img')[0]['alt']
            source = book('img')[0]['src']
            cover_link = "http://books.toscrape.com"+source[2:]

            price_cell = sheet3.cell(row = i3, column = 3)
            price_cell.value = money

            title_cell = sheet3.cell(row = i3, column = 2)
            title_cell.value = title

            cover_cell = sheet3.cell(row = i3, column = 4)
            cover_cell.value = cover_link
            i3+=1

        # 4 stars books
        if len(book.select('.star-rating.Four')) != 0:
            sheet4 = file['4 stars']

            money = book.select('.price_color')[0].text[1:]
            title = book('img')[0]['alt']
            source = book('img')[0]['src']
            cover_link = "http://books.toscrape.com"+source[2:]

            price_cell = sheet4.cell(row = i4, column = 3)
            price_cell.value = money

            title_cell = sheet4.cell(row = i4, column = 2)
            title_cell.value = title

            cover_cell = sheet4.cell(row = i4, column = 4)
            cover_cell.value = cover_link
            i4+=1

        # 5 stars books
        elif len(book.select('.star-rating.Five')) != 0:
            sheet5 = file['5 stars']

            money = book.select('.price_color')[0].text[1:]
            title = book('img')[0]['alt']
            source = book('img')[0]['src']
            cover_link = "http://books.toscrape.com"+source[2:]

            price_cell = sheet5.cell(row = i5, column = 3)
            price_cell.value = money

            title_cell = sheet5.cell(row = i5, column = 2)
            title_cell.value = title

            cover_cell = sheet5.cell(row = i5, column = 4)
            cover_cell.value = cover_link
            i5+=1


file.save('example.xlsx')



"""
# get title
url = "http://books.toscrape.com/catalogue/page-1.html"

re = requests.get(url)
soup = bs4.BeautifulSoup(re.text, "lxml")

books_images = soup.select('.thumbnail') # get title & cover_link
prices  = soup.select('.price_color') # prices

file = openpyxl.load_workbook('test.xlsx')

sheet_obj = file.active

for i, price in enumerate(prices):
    i+=3
    money = price.text[1:]
    price_cell = sheet_obj.cell(row = i, column = 3)
    price_cell.value = money

for n, book in enumerate(books_images):
    n+=3
    title = book['alt']
    # cover link
    source = book['src'] 
    cover_link = "http://books.toscrape.com"+source[2:]

    title_cell = sheet_obj.cell(row = n, column = 2)
    title_cell.value = title

    cover_link_cell = sheet_obj.cell(row = n, column = 4)
    cover_link_cell.value = cover_link


file.save('example.xlsx')

"""

"""
sheets = ['2 stars', '3 stars', '4 stars']

file = openpyxl.load_workbook('template.xlsx')
sheet = file.active


for n, item in enumerate(sheets):
    n+=5
    sheet = file[item]
    sheet.cell(row = 1, column = 1).value = n


file.save('example.xlsx')
"""

"""
sheet2 = file['2 stars']
        money = book.select('.price_color')[0].text[1:]
        price_cell = sheet2.cell(row = i, column = 3)
        price_cell.value = money
"""
















  













    





